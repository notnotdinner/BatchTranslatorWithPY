import requests
import json
import openpyxl
from types import SimpleNamespace as Namespace
from translatetool.LanguagePostfix import LanguagePostfix
from translatetool.ms.ms_translate_result import *
from translatetool.ms.ms_language import MSSupportLanguage
import os

AZURE_TRANSLATE_SERVICE_ENDPOINT = "https://api.cognitive.microsofttranslator.com"

ms_token = ""


def translate_words(words: list, languages: list):

    to_lans = ""
    result = dict()
    for lan_postfix in languages:
        result[lan_postfix] = []
        lan_code = LanguagePostfix.get_lan_code_by_postfix(lan_postfix)
        if len(to_lans) > 0:
            to_lans += ","
        to_lans += lan_code

    request_url = AZURE_TRANSLATE_SERVICE_ENDPOINT + "/translate?api-version=3.0&to=" + to_lans

    my_headers = {"Ocp-Apim-Subscription-Key": ms_token, "Content-type": "application/json"}

    request_param = []
    for word in words:
        request_param.append({"Text": word})

    translate_response = requests.post(request_url, json=request_param, headers=my_headers)

    print(translate_response.text)

    ms_translate_results = json.loads(translate_response.text, object_hook=lambda d: Namespace(**d))

    for ms_translate_result in ms_translate_results:

        for translate_item in ms_translate_result.translations:
            lan_postfix = LanguagePostfix.get_post_fix_by_ms_lan_code(translate_item.to)
            result_array = result[lan_postfix]
            result_array.append(translate_item.text)
            print(translate_item.to, translate_item.text)

    return result


def input_token():
    token = input("Please input Azure token:")
    token_file = open("ms_token.txt", "w+")
    token_file.write(ms_token)
    token_file.close()
    return token


if __name__ == '__main__':

    current_path = os.getcwd()
    token_file_name = os.path.join(current_path, "ms_token.txt")

    if not os.path.exists(token_file_name):
        token = input_token()
    else:
        ms_token_file = open(token_file_name, "r")
        token = ms_token_file.read()
        ms_token_file.close()
        if token is None or len(token) == 0:
            token = input_token()

    ms_token = token
    excel_file_path = input("Please input the file to translate:")
    excel_file_path.rstrip()
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.worksheets[0]

    col_to_translate = -1
    target_title = ""
    for col in range(1, sheet.max_column+1):
        font_color = sheet.cell(1, col).font.color
        if font_color.rgb == "FFFF0000":
            col_to_translate = col
            target_title = str(sheet.cell(1, col).value)
            break

    languages_to_translate = []
    language_indexes = []
    for col in range(1, sheet.max_column+1):
        cell = sheet.cell(1, col)
        if cell is not None:
            cell_title = str(cell.value)
            if cell_title.startswith(target_title) and len(cell_title) > len(target_title):
                postfix = cell_title[len(target_title) + 1:]
                if postfix in LanguagePostfix.get_language_postfix_list():
                    languages_to_translate.append(postfix)
                    language_indexes.append(col)

    if col_to_translate > 0:
        words_to_translate = []
        row_indexes_to_translate = []
        for row in range(2, sheet.max_row+1):
            cell = sheet.cell(row, col_to_translate)
            if cell.value is not None:
                words_to_translate.append(str(cell.value))
                row_indexes_to_translate.append(row)

        translate_results = translate_words(words_to_translate, languages_to_translate)

        for idx in range(0, len(language_indexes)):
            col = language_indexes[idx]
            postfix = languages_to_translate[idx]
            translated_texts = translate_results[postfix]
            if translated_texts is not None:
                row = 2
                for text in translated_texts:
                    sheet.cell(row, col, text)
                    row += 1

        workbook.save(excel_file_path)
        workbook.close()

    # request_url = AZURE_TRANSLATE_SERVICE_ENDPOINT + "/translate?api-version=3.0&to=th,en,ja,zh-Hans,ru,ko"
    #
    # my_headers = {"Ocp-Apim-Subscription-Key": AZURE_TOKEN, "Content-type": "application/json"}
    #
    # translate_response = requests.post(request_url, json=request_param, headers=my_headers)
    #
    # print(translate_response.text)
    #
    # ms_translate_results = json.loads(translate_response.text, object_hook=lambda d: Namespace(**d))
    #
    # ms_translate_result: MSTranslateResult
    #
    # for ms_translate_result in ms_translate_results:
    #     print(ms_translate_result.detectedLanguage.language)
    #     for translate_item in ms_translate_result.translations:
    #         print(translate_item.to)



